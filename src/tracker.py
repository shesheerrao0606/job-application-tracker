"""Tracker spreadsheet management.

Schema (one sheet, "Applications"):
    A  Date Applied      (YYYY-MM-DD)
    B  Company
    C  Role
    D  Status            (Applied | Interview | Offer | Rejected)
    E  Last Update       (YYYY-MM-DD)
    F  Source Account    (which gmail address)
    G  Notes             (auto-appended classifier reasons + snippets)
    H  Last Message ID   (most recent gmail msg id touching this row)
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger("tracker")

HEADERS = [
    "Date Applied",
    "Company",
    "Role",
    "Status",
    "Last Update",
    "Source Account",
    "Notes",
    "Last Message ID",
]

STATUS_FROM_CATEGORY = {
    "applied": "Applied",
    "rejected": "Rejected",
    "interview": "Interview",
    "offer": "Offer",
}

# Status priority — only "upgrade" rows. Don't let a stray "Applied" classification
# overwrite an existing "Offer".
STATUS_PRIORITY = {"Applied": 0, "Rejected": 1, "Interview": 2, "Offer": 3}


def reconcile_into_tracker(classified: list[dict], path: Path) -> dict[str, int]:
    """Insert/update rows; return {new, updated, skipped} counts."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = _load_or_create(path)
    ws = wb["Applications"]

    rows = _read_rows(ws)
    counts = {"new": 0, "updated": 0, "skipped": 0}

    # Sort: process "applied" first so rejection/interview can match against rows
    # we just created (e.g. rejection arrives same day as confirmation)
    order = {"applied": 0, "interview": 1, "offer": 1, "rejected": 2}
    classified_sorted = sorted(classified, key=lambda c: order.get(c["category"], 9))

    for c in classified_sorted:
        new_status = STATUS_FROM_CATEGORY.get(c["category"])
        if not new_status:
            counts["skipped"] += 1
            continue
        date_iso = _email_date_iso(c["date"])

        if c["category"] == "applied":
            # Skip if there's already a recent row for the same (company, role,
            # account) — likely a duplicate "we received your application"
            # auto-reply, or a stray confirmation arriving after the row has
            # already been advanced to Interview/Offer.
            dup_idx = _find_recent_duplicate(rows, c, days=30)
            if dup_idx is not None:
                rows[dup_idx]["Notes"] = (
                    rows[dup_idx]["Notes"] + "\n"
                    + _format_note(c, "Applied (duplicate, ignored)")
                ).strip()
                counts["skipped"] += 1
                continue

            row = {
                "Date Applied": date_iso,
                "Company": c["company"] or "(unknown)",
                "Role": c["role"] or "(unknown)",
                "Status": "Applied",
                "Last Update": date_iso,
                "Source Account": c["account"],
                "Notes": _format_note(c, "Applied"),
                "Last Message ID": c["id"],
            }
            rows.append(row)
            counts["new"] += 1
            continue

        # rejection / interview / offer — try to match an existing row
        match_idx = _find_match(rows, c)
        if match_idx is None:
            # Orphan update: still record it so nothing is silently dropped
            rows.append({
                "Date Applied": "",
                "Company": c["company"] or "(unknown)",
                "Role": c["role"] or "(unknown)",
                "Status": new_status,
                "Last Update": date_iso,
                "Source Account": c["account"],
                "Notes": "[orphan — no matching application found] "
                         + _format_note(c, new_status),
                "Last Message ID": c["id"],
            })
            counts["new"] += 1
            continue

        existing = rows[match_idx]
        # Only upgrade status if the new one is "later" in the funnel
        if STATUS_PRIORITY[new_status] >= STATUS_PRIORITY.get(existing["Status"], 0):
            existing["Status"] = new_status
            existing["Last Update"] = date_iso
            existing["Notes"] = (existing["Notes"] + "\n" + _format_note(c, new_status)).strip()
            existing["Last Message ID"] = c["id"]
            counts["updated"] += 1
        else:
            counts["skipped"] += 1

    _write_rows(ws, rows)
    _write_stats_sheet(wb, rows)
    wb.save(path)
    return counts


# --- helpers ------------------------------------------------------------------

def _load_or_create(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E78")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    widths = {"A": 13, "B": 22, "C": 28, "D": 12, "E": 13, "F": 24, "G": 50, "H": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    return wb


def _read_rows(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        rows.append({HEADERS[i]: (r[i] if i < len(r) else "") or "" for i in range(len(HEADERS))})
    return rows


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    # Clear everything below the header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=i, column=j, value=row.get(h, ""))
            cell.font = Font(name="Arial")
            if h == "Notes":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if h == "Status":
                cell.fill = _status_fill(str(row.get(h, "")))


def _write_stats_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    """Recreate the 'Stats' sheet with current numbers + a status-distribution chart."""
    # Lazy import — keeps this file's surface clean
    from stats import _compute  # type: ignore
    from accounts import display_name

    s = _compute(rows)

    # Drop and recreate so we never accumulate stale rows
    if "Stats" in wb.sheetnames:
        del wb["Stats"]
    ws = wb.create_sheet("Stats")

    # Column widths
    for col, w in {"A": 28, "B": 14, "C": 30, "D": 14}.items():
        ws.column_dimensions[col].width = w

    bold = Font(name="Arial", bold=True)
    header_fill = PatternFill("solid", start_color="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")

    def section(row: int, title: str) -> int:
        c = ws.cell(row=row, column=1, value=title)
        c.font = header_font
        c.fill = header_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        return row + 1

    r = 1
    r = section(r, "Overview")
    overview = [
        ("Total applications", s["total"]),
        ("This week", s["this_week"]),
        ("This month", s["this_month"]),
        ("Mature applications (>21d old)", s["mature_count"]),
        ("Avg response time (days)",
         s["avg_response_days"] if s["avg_response_days"] is not None else "—"),
        ("Last updated",
         datetime.now(timezone.utc).date().isoformat()),
    ]
    for label, val in overview:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    r = section(r, "Rates (% of mature applications)")
    rates = [
        ("Response rate", s["response_rate"]),
        ("Interview rate", s["interview_rate"]),
        ("Rejection rate", s["rejection_rate"]),
    ]
    for label, val in rates:
        ws.cell(row=r, column=1, value=label).font = bold
        c = ws.cell(row=r, column=2, value=val / 100)
        c.number_format = "0.0%"
        r += 1

    r += 1
    r = section(r, "Status distribution")
    dist_start = r
    for status, count in sorted(s["status_dist"].items(), key=lambda kv: -kv[1]):
        ws.cell(row=r, column=1, value=status).font = bold
        ws.cell(row=r, column=2, value=count)
        r += 1
    dist_end = r - 1

    # Add a bar chart for the distribution if we have data
    if dist_end >= dist_start:
        try:
            from openpyxl.chart import BarChart, Reference
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Status distribution"
            chart.legend = None
            chart.height = 7
            chart.width = 12
            data = Reference(ws, min_col=2, min_row=dist_start, max_row=dist_end)
            cats = Reference(ws, min_col=1, min_row=dist_start, max_row=dist_end)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            ws.add_chart(chart, "D2")
        except Exception as e:
            log.warning("could not add chart: %s", e)

    r += 1
    r = section(r, "Top companies")
    if not s["top_companies"]:
        ws.cell(row=r, column=1, value="(none)").font = Font(name="Arial", italic=True, color="888888")
        r += 1
    else:
        for company, count in s["top_companies"]:
            ws.cell(row=r, column=1, value=company).font = bold
            ws.cell(row=r, column=2, value=count)
            r += 1

    r += 1
    r = section(r, "By Gmail account")
    # column headers
    for j, h in enumerate(["Account", "Applied", "Interview/Offer", "Rejected"], start=1):
        cell = ws.cell(row=r, column=j, value=h)
        cell.font = bold
        cell.fill = PatternFill("solid", start_color="DDEBF7")
    r += 1
    if not s["by_account"]:
        ws.cell(row=r, column=1, value="(none)").font = Font(name="Arial", italic=True, color="888888")
        r += 1
    else:
        for acct, d in sorted(s["by_account"].items(), key=lambda kv: -kv[1]["applied"]):
            ws.cell(row=r, column=1, value=display_name(acct))
            ws.cell(row=r, column=2, value=d["applied"])
            ws.cell(row=r, column=3, value=d["interview"] + d["offer"])
            ws.cell(row=r, column=4, value=d["rejected"])
            r += 1


def _status_fill(status: str) -> PatternFill:
    colors = {
        "Applied": "DDEBF7",     # light blue
        "Interview": "FFF2CC",   # light yellow
        "Offer": "C6EFCE",       # light green
        "Rejected": "F8CBAD",    # light red/orange
    }
    color = colors.get(status, "FFFFFF")
    return PatternFill("solid", start_color=color)


def _find_recent_duplicate(rows: list[dict], c: dict, *, days: int) -> int | None:
    """Find a row with the same (company, role, account) updated within `days`.

    Used to suppress duplicate 'applied' confirmations.
    """
    from datetime import date, timedelta
    company = (c["company"] or "").strip().lower()
    role = (c["role"] or "").strip().lower()
    account = c["account"]
    if not company or not role:
        return None
    cutoff = date.today() - timedelta(days=days)
    for idx, r in enumerate(rows):
        if str(r.get("Company", "")).strip().lower() != company:
            continue
        if str(r.get("Role", "")).strip().lower() != role:
            continue
        if r.get("Source Account") != account:
            continue
        last = str(r.get("Last Update", ""))
        try:
            last_d = datetime.fromisoformat(last).date()
        except ValueError:
            continue
        if last_d >= cutoff:
            return idx
    return None


def _find_match(rows: list[dict], c: dict) -> int | None:
    """Match by (company, account) — prefer exact role match, then most recent."""
    company = (c["company"] or "").strip().lower()
    role = (c["role"] or "").strip().lower()
    account = c["account"]
    if not company:
        return None

    candidates: list[tuple[int, dict]] = []
    for idx, r in enumerate(rows):
        if str(r.get("Company", "")).strip().lower() != company:
            continue
        if r.get("Source Account") and r["Source Account"] != account:
            continue
        candidates.append((idx, r))

    if not candidates:
        return None

    # Prefer rows whose role matches; otherwise most recent Last Update
    if role:
        exact = [(i, r) for i, r in candidates if str(r.get("Role", "")).strip().lower() == role]
        if exact:
            return exact[-1][0]

    candidates.sort(key=lambda ir: str(ir[1].get("Last Update", "")), reverse=True)
    return candidates[0][0]


def _email_date_iso(raw: str) -> str:
    if not raw:
        return datetime.now(timezone.utc).date().isoformat()
    try:
        return parsedate_to_datetime(raw).date().isoformat()
    except Exception:
        return datetime.now(timezone.utc).date().isoformat()


def _format_note(c: dict, status: str) -> str:
    today = datetime.now(timezone.utc).date().isoformat()
    return f"[{today} {status}] {c.get('reason', '')} (msg:{c['id']})"