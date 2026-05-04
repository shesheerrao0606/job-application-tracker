"""Statistics computed from the tracker state.

Reads the Applications sheet and computes stats used in both the digest
email and the Stats sheet inside tracker.xlsx.

Design notes:
- "Response rate" is computed against applications old enough to have had
  a fair chance at a response (default: 21 days). Otherwise last week's
  applications drag the rate down artificially.
- "Avg response time" only counts applications where Last Update > Date
  Applied (i.e., something actually came back).
- Per-account breakdown shows applications + responses per inbox.
- "Top companies" handles dupes (multiple roles at same company) by counting
  rows, not unique companies.
"""
from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

RESPONSE_WINDOW_DAYS = 21


def compute_stats(tracker_path: Path) -> dict[str, Any]:
    """Read tracker.xlsx and return a stats dict."""
    if not tracker_path.exists():
        return _empty_stats()

    wb = load_workbook(tracker_path, data_only=True)
    if "Applications" not in wb.sheetnames:
        return _empty_stats()

    ws = wb["Applications"]
    rows = _read_rows(ws)
    return _compute(rows)


def _read_rows(ws) -> list[dict[str, Any]]:
    headers = [c.value for c in ws[1]]
    out: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        out.append({headers[i]: (r[i] if i < len(r) else "") or "" for i in range(len(headers))})
    return out


def _compute(rows: list[dict[str, Any]]) -> dict[str, Any]:
    today = date.today()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    response_cutoff = today - timedelta(days=RESPONSE_WINDOW_DAYS)

    # Filter to rows that represent real applications (have a Date Applied)
    applied_rows = [r for r in rows if _parse_date(r.get("Date Applied"))]

    total = len(applied_rows)
    this_week = sum(1 for r in applied_rows if _parse_date(r["Date Applied"]) >= week_ago)
    this_month = sum(1 for r in applied_rows if _parse_date(r["Date Applied"]) >= month_ago)

    # Mature applications = applied at least RESPONSE_WINDOW_DAYS ago
    mature = [r for r in applied_rows if _parse_date(r["Date Applied"]) <= response_cutoff]
    mature_count = len(mature)

    responded = [r for r in mature if str(r.get("Status", "")).lower() != "applied"]
    interviewed = [r for r in mature if str(r.get("Status", "")).lower() in ("interview", "offer")]
    rejected = [r for r in mature if str(r.get("Status", "")).lower() == "rejected"]

    response_rate = (len(responded) / mature_count * 100) if mature_count else 0.0
    interview_rate = (len(interviewed) / mature_count * 100) if mature_count else 0.0
    rejection_rate = (len(rejected) / mature_count * 100) if mature_count else 0.0

    # Avg response time = days from Date Applied → Last Update, for any
    # row whose status moved past Applied
    response_times = []
    for r in applied_rows:
        if str(r.get("Status", "")).lower() == "applied":
            continue
        applied_d = _parse_date(r["Date Applied"])
        update_d = _parse_date(r.get("Last Update"))
        if applied_d and update_d and update_d >= applied_d:
            response_times.append((update_d - applied_d).days)
    avg_response_days = (sum(response_times) / len(response_times)) if response_times else None

    # Top companies (rows, not unique companies — Acme x3 means 3 attempts)
    top_companies = Counter(
        str(r.get("Company", "")).strip() for r in applied_rows
        if str(r.get("Company", "")).strip() and str(r.get("Company", "")).strip() != "(unknown)"
    ).most_common(5)

    # Per-account breakdown
    by_account: dict[str, dict[str, int]] = defaultdict(lambda: {
        "applied": 0, "rejected": 0, "interview": 0, "offer": 0
    })
    for r in applied_rows:
        acct = str(r.get("Source Account", "")).strip() or "(unknown)"
        status = str(r.get("Status", "")).lower()
        by_account[acct]["applied"] += 1
        if status in by_account[acct]:
            if status != "applied":
                by_account[acct][status] += 1

    # Status distribution across the whole tracker (including orphan rejections)
    status_dist = Counter(str(r.get("Status", "")).strip() or "Unknown" for r in rows)

    return {
        "total": total,
        "this_week": this_week,
        "this_month": this_month,
        "mature_count": mature_count,
        "response_rate": round(response_rate, 1),
        "interview_rate": round(interview_rate, 1),
        "rejection_rate": round(rejection_rate, 1),
        "avg_response_days": round(avg_response_days, 1) if avg_response_days is not None else None,
        "top_companies": top_companies,
        "by_account": dict(by_account),
        "status_dist": dict(status_dist),
        "response_window_days": RESPONSE_WINDOW_DAYS,
    }


def _empty_stats() -> dict[str, Any]:
    return {
        "total": 0, "this_week": 0, "this_month": 0, "mature_count": 0,
        "response_rate": 0.0, "interview_rate": 0.0, "rejection_rate": 0.0,
        "avg_response_days": None, "top_companies": [], "by_account": {},
        "status_dist": {}, "response_window_days": RESPONSE_WINDOW_DAYS,
    }


def _parse_date(v: Any) -> date | None:
    if not v:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return None


# --- ASCII bar rendering for the digest email -------------------------------

def hbar(value: float, scale: float, *, width: int = 24) -> str:
    """Solid-block bar of length proportional to value/scale."""
    if scale <= 0:
        return ""
    filled = max(0, min(width, round(width * value / scale)))
    return "█" * filled + "░" * (width - filled)